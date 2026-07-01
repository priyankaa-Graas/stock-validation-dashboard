import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Color constants
HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Indigo
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

TRUE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
TRUE_FONT = Font(name="Segoe UI", size=10, color="375623")

FALSE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Orange/Red
FALSE_FONT = Font(name="Segoe UI", size=10, color="C65911")

ALL_GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
ALL_GOOD_FONT = Font(name="Segoe UI", size=10, color="006100")

WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Soft Yellow
WARN_FONT = Font(name="Segoe UI", size=10, color="9C6500")

ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Soft Red
ERROR_FONT = Font(name="Segoe UI", size=10, color="9C0006")

DEFAULT_FONT = Font(name="Segoe UI", size=10)
BORDER_THIN = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

def autofit_columns(ws):
    """
    Autofits worksheet column widths.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Handle linebreaks or formatting by converting to string
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Apply padding and set width
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def style_ws_headers(ws):
    """
    Applies standard styling to the first row headers.
    """
    ws.row_dimensions[1].height = 26
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def generate_status_report(df):
    """
    Generates status validation report with styling.
    """
    output = io.BytesIO()
    
    # Columns: SKU, Product ID, Marketplace Status, TC Status, Marketplace Stock, TC Stock, Reserved Stock, Max, Status Check, Stock Check, Recommendation
    export_df = df[[
        'sku', 'product_id', 'marketplace_status', 'tc_status', 
        'marketplace_stock', 'tc_stock', 'reserved_stock', 'max', 
        'status_check', 'stock_check', 'recommendation'
    ]].copy()
    
    # Rename for professional look
    export_df.columns = [
        'SKU', 'Product ID', 'Marketplace Status', 'TC Status', 
        'Marketplace Stock', 'TC Stock', 'Reserved Stock', 'Max (0)', 
        'Status Check', 'Stock Check', 'Recommendation'
    ]
    
    # Create workbook
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Status Validation')
        workbook = writer.book
        worksheet = writer.sheets['Status Validation']
        
        # Style header
        style_ws_headers(worksheet)
        
        # Style rows
        for r_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[r_idx].height = 20
            
            # Read columns we need to color code (1-based index)
            # SKU: col 1, Product ID: col 2, etc.
            # Status Check: col 9, Stock Check: col 10, Recommendation: col 11
            status_cell = worksheet.cell(row=r_idx, column=9)
            stock_cell = worksheet.cell(row=r_idx, column=10)
            rec_cell = worksheet.cell(row=r_idx, column=11)
            
            # Format status check
            if status_cell.value == True or str(status_cell.value).strip().upper() == "TRUE":
                status_cell.fill = TRUE_FILL
                status_cell.font = TRUE_FONT
            else:
                status_cell.fill = FALSE_FILL
                status_cell.font = FALSE_FONT
                
            # Format stock check
            if stock_cell.value == True or str(stock_cell.value).strip().upper() == "TRUE":
                stock_cell.fill = TRUE_FILL
                stock_cell.font = TRUE_FONT
            else:
                stock_cell.fill = FALSE_FILL
                stock_cell.font = FALSE_FONT
                
            # Format recommendation
            rec_val = str(rec_cell.value or '').strip()
            if rec_val == "All Good":
                rec_cell.fill = ALL_GOOD_FILL
                rec_cell.font = ALL_GOOD_FONT
            elif "Change to" in rec_val:
                rec_cell.fill = WARN_FILL
                rec_cell.font = WARN_FONT
            else:
                rec_cell.fill = ERROR_FILL
                rec_cell.font = ERROR_FONT
                
            # Basic styling for all cells in the row
            for c_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                if c_idx not in [9, 10, 11]: # Don't overwrite colored columns
                    cell.font = DEFAULT_FONT
                cell.border = BORDER_THIN
                cell.alignment = Alignment(vertical="center", horizontal="left" if c_idx in [1, 2, 11] else "center")
                
        autofit_columns(worksheet)
        
    return output.getvalue()

def generate_pid_report(df):
    """
    Generates Product ID validation report with styling.
    """
    output = io.BytesIO()
    
    if df.empty:
        # Create empty styled sheet
        df = pd.DataFrame(columns=['Product ID', 'Marketplace Stock', 'TC Stock', 'Stock Check', 'Recommendation'])
        
    export_df = df.copy()
    export_df.columns = ['Product ID', 'Marketplace Stock (Consolidated)', 'TC Stock (Consolidated)', 'Stock Check', 'Recommendation']
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='PID Validation')
        workbook = writer.book
        worksheet = writer.sheets['PID Validation']
        
        style_ws_headers(worksheet)
        
        for r_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[r_idx].height = 20
            
            stock_check_cell = worksheet.cell(row=r_idx, column=4)
            rec_cell = worksheet.cell(row=r_idx, column=5)
            
            if stock_check_cell.value == True or str(stock_check_cell.value).strip().upper() == "TRUE":
                stock_check_cell.fill = TRUE_FILL
                stock_check_cell.font = TRUE_FONT
            else:
                stock_check_cell.fill = FALSE_FILL
                stock_check_cell.font = FALSE_FONT
                
            rec_val = str(rec_cell.value or '').strip()
            if rec_val == "All Good":
                rec_cell.fill = ALL_GOOD_FILL
                rec_cell.font = ALL_GOOD_FONT
            else:
                rec_cell.fill = ERROR_FILL
                rec_cell.font = ERROR_FONT
                
            for c_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                if c_idx not in [4, 5]:
                    cell.font = DEFAULT_FONT
                cell.border = BORDER_THIN
                cell.alignment = Alignment(vertical="center", horizontal="left" if c_idx in [1, 5] else "center")
                
        autofit_columns(worksheet)
        
    return output.getvalue()

def generate_bundle_report(df):
    """
    Generates Bundle SKU report with styling.
    """
    output = io.BytesIO()
    
    if df.empty:
        df = pd.DataFrame(columns=['Bundle SKU', 'Marketplace Stock', 'Calculated TC Stock', 'Type', 'Components Summary'])
        
    export_df = df.copy()
    export_df.columns = ['Bundle SKU', 'Marketplace Stock', 'Calculated TC Stock', 'Bundle Type', 'Components & Formulas']
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Bundle Validation')
        workbook = writer.book
        worksheet = writer.sheets['Bundle Validation']
        
        style_ws_headers(worksheet)
        
        for r_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[r_idx].height = 20
            
            for c_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                cell.font = DEFAULT_FONT
                cell.border = BORDER_THIN
                cell.alignment = Alignment(vertical="center", horizontal="left" if c_idx in [1, 5] else "center")
                
        autofit_columns(worksheet)
        
    return output.getvalue()

def generate_error_report(errors_list):
    """
    Generates Error Report with styling.
    """
    output = io.BytesIO()
    
    df = pd.DataFrame(errors_list)
    if df.empty:
        df = pd.DataFrame(columns=['file', 'row', 'sku', 'error'])
        
    export_df = df.copy()
    export_df.columns = ['File Name', 'Row Number', 'SKU / Product ID', 'Error Description']
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Errors & Warnings')
        workbook = writer.book
        worksheet = writer.sheets['Errors & Warnings']
        
        style_ws_headers(worksheet)
        
        for r_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[r_idx].height = 20
            
            error_desc_cell = worksheet.cell(row=r_idx, column=4)
            error_desc_cell.fill = ERROR_FILL
            error_desc_cell.font = ERROR_FONT
            
            for c_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                if c_idx != 4:
                    cell.font = DEFAULT_FONT
                cell.border = BORDER_THIN
                cell.alignment = Alignment(vertical="center", horizontal="left" if c_idx in [1, 3, 4] else "center")
                
        autofit_columns(worksheet)
        
    return output.getvalue()
